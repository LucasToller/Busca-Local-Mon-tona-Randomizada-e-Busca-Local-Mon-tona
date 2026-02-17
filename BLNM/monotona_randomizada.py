import os
import random
import time
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from collections import defaultdict
from openpyxl.utils import get_column_letter

# ============================================================
# BLNM = Busca Local Monótona Randomizada
# - Com probabilidade alpha: passo aleatório (caminhada aleatória)
# - Caso contrário: melhor melhora (best improvement)
#
# Monotonia: medida pelo BEST-SO-FAR (melhor valor encontrado)
# Parada: 1000 iterações sem melhorar o best-so-far
#
# Saídas geradas em: BLNM\Resultados\
#   - resultados_blnm.txt
#   - resultados_blnm.xlsx (com 2 abas: resultados + resumo)
# ============================================================


def construir_solucao_inicial(n, m, tempos):
    """Solução inicial aleatória + cargas."""
    sol = [random.randrange(m) for _ in range(n)]
    cargas = [0] * m
    for i, maq in enumerate(sol):
        cargas[maq] += tempos[i]
    return sol, cargas


def makespan(cargas):
    return max(cargas)


def passo_aleatorio(sol, cargas, tempos, m):
    """Move uma tarefa para outra máquina aleatória e atualiza cargas."""
    n = len(sol)
    tarefa = random.randrange(n)
    origem = sol[tarefa]

    destino = random.randrange(m)
    while destino == origem:
        destino = random.randrange(m)

    p = tempos[tarefa]
    sol[tarefa] = destino
    cargas[origem] -= p
    cargas[destino] += p


def top3_cargas(cargas):
    """
    Retorna uma lista com até os 3 maiores pares (carga, idx_maquina),
    em ordem decrescente.
    """
    pares = [(c, i) for i, c in enumerate(cargas)]
    pares.sort(reverse=True)
    return pares[:3]


def maior_excluindo(top3, a, b):
    """
    Retorna a maior carga entre as máquinas, excluindo a e b, usando apenas o top3.
    """
    for c, i in top3:
        if i != a and i != b:
            return c
    return 0


def avaliar_melhor_melhora(sol, cargas, tempos, m):
    """
    Best improvement (melhor melhora) em toda a vizinhança.

    Otimização:
    novo makespan = max(nova_carga_origem, nova_carga_destino, maior_carga_das_outras)
    onde "maior_carga_das_outras" é obtida por top3, evitando loop em m.
    """
    valor_atual = makespan(cargas)
    n = len(tempos)

    melhor_valor = valor_atual
    melhor_tarefa = None
    melhor_origem = None
    melhor_destino = None

    t3 = top3_cargas(cargas)

    for tarefa in range(n):
        origem = sol[tarefa]
        p = tempos[tarefa]

        for destino in range(m):
            if destino == origem:
                continue

            nova_origem = cargas[origem] - p
            nova_dest = cargas[destino] + p

            outras = maior_excluindo(t3, origem, destino)
            novo_ms = max(nova_origem, nova_dest, outras)

            if novo_ms < melhor_valor:
                melhor_valor = novo_ms
                melhor_tarefa = tarefa
                melhor_origem = origem
                melhor_destino = destino

    return melhor_tarefa, melhor_origem, melhor_destino, melhor_valor


def blnm_monotona_randomizada(tempos, m, alpha, max_sem_melhora=1000):
    """
    Busca Local Monótona Randomizada:
    - alpha: frequência de caminhada aleatória
    - best-so-far é o que conta para o contador sem melhora
    """
    n = len(tempos)
    sol, cargas = construir_solucao_inicial(n, m, tempos)

    best = makespan(cargas)
    sem_melhora = 0
    it = 0
    inicio = time.time()

    while sem_melhora < max_sem_melhora:
        it += 1

        if random.random() < alpha:
            passo_aleatorio(sol, cargas, tempos, m)
            valor_atual = makespan(cargas)
        else:
            tarefa, origem, destino, novo_valor = avaliar_melhor_melhora(sol, cargas, tempos, m)

            if tarefa is not None:
                p = tempos[tarefa]
                sol[tarefa] = destino
                cargas[origem] -= p
                cargas[destino] += p
                valor_atual = novo_valor
            else:
                valor_atual = best

        if valor_atual < best:
            best = valor_atual
            sem_melhora = 0
        else:
            sem_melhora += 1

    tempo_exec = time.time() - inicio
    return best, it, tempo_exec


def exportar_txt(caminho, linhas):
    with open(caminho, "w", encoding="utf-8") as f:
        f.write("heuristica,n,m,replicacao,tempo,iteracoes,valor,parametro\n")
        for heur, n, m, rep, tempo, it, val, alpha in linhas:
            f.write(f"{heur},{n},{m},{rep},{tempo:.4f},{it},{val},{alpha:.1f}\n")


def formatar_tempo_min_seg(segundos):
    """Converte segundos (float) para string 'Xm Ys'."""
    total = int(round(segundos))
    m = total // 60
    s = total % 60
    return f"{m}m {s}s"


def estilizar_cabecalho(ws, num_cols):
    """Cabeçalho cinza claro, negrito, centralizado e congelado."""
    cinza_claro = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    fonte = Font(bold=True)
    alinhamento = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"

    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = cinza_claro
        cell.font = fonte
        cell.alignment = alinhamento

    ws.auto_filter.ref = ws.dimensions


def ajustar_largura_colunas(ws):
    """Ajusta largura das colunas com base no maior conteúdo."""
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0

        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = cell.value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))

        ws.column_dimensions[col_letter].width = min(max_len + 2, 35)


def criar_aba_resumo(wb, linhas, tempo_total_script, config):
    """
    Cria uma segunda aba 'resumo' com:
    - tempo total do script
    - contagens e parâmetros do experimento
    - estatísticas rápidas
    - médias por alpha (opcional e útil)
    """
    ws = wb.create_sheet("resumo")

    # Estilo de título e chave/valor
    titulo_font = Font(bold=True, size=13)
    key_font = Font(bold=True)
    cinza_claro = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    # Título
    ws["A1"] = "Resumo de Execução - BLNM (Monótona Randomizada)"
    ws["A1"].font = titulo_font
    ws.merge_cells("A1:D1") 
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # Linha 2 como “cabeçalho” da seção
    ws["A2"] = "Item"
    ws["B2"] = "Valor"
    ws["A2"].font = key_font
    ws["B2"].font = key_font
    ws["A2"].fill = cinza_claro
    ws["B2"].fill = cinza_claro
    ws["A2"].alignment = center
    ws["B2"].alignment = center

    # Dados básicos
    total_registros = len(linhas)
    tempos = [t for (_, _, _, _, t, _, _, _) in linhas]
    iteracoes = [it for (_, _, _, _, _, it, _, _) in linhas]
    valores = [v for (_, _, _, _, _, _, v, _) in linhas]

    ws.append(["Tempo total do script", formatar_tempo_min_seg(tempo_total_script)])
    ws.append(["Tempo total do script (s)", f"{tempo_total_script:.2f}"])

    ws.append(["Total de registros", total_registros])
    ws.append(["Registros esperados", config["esperado_registros"]])
    ws.append(["m utilizados", str(config["maquinas"])])
    ws.append(["r utilizados (n = m*r)", str(config["rs"])])
    ws.append(["Repetições", config["repeticoes"]])
    ws.append(["Alphas", str(config["alphas"])])
    ws.append(["Parada (sem melhora)", config["max_sem_melhora"]])

    # Stats rápidas
    ws.append(["Tempo médio por execução (s)", f"{(sum(tempos) / total_registros):.4f}"])
    ws.append(["Tempo mínimo por execução (s)", f"{min(tempos):.4f}"])
    ws.append(["Tempo máximo por execução (s)", f"{max(tempos):.4f}"])

    ws.append(["Iterações médias", int(sum(iteracoes) / total_registros)])
    ws.append(["Iterações mínimas", min(iteracoes)])
    ws.append(["Iterações máximas", max(iteracoes)])

    ws.append(["Melhor valor (menor makespan)", min(valores)])
    ws.append(["Pior valor (maior makespan)", max(valores)])

    # Ajusta estilo na coluna A (chaves)
    for row in range(3, ws.max_row + 1):
        ws.cell(row=row, column=1).font = key_font

    # Seção extra: médias por alpha
    ws["A20"] = "Médias por alpha (parametro)"
    ws["A20"].font = titulo_font
    ws.merge_cells("A20:D20")
    ws["A20"].alignment = Alignment(horizontal="center", vertical="center")

    ws.append(["alpha", "valor médio", "tempo médio (s)", "iterações médias"])
    header_row = ws.max_row
    for col in range(1, 5):
        c = ws.cell(row=header_row, column=col)
        c.fill = cinza_claro
        c.font = key_font
        c.alignment = center

    por_alpha = defaultdict(lambda: {"val": 0, "tempo": 0, "it": 0, "count": 0})
    for (_, _, _, _, tempo, it, val, alpha) in linhas:
        a = round(float(alpha), 1)
        por_alpha[a]["val"] += val
        por_alpha[a]["tempo"] += tempo
        por_alpha[a]["it"] += it
        por_alpha[a]["count"] += 1

    for a in sorted(por_alpha.keys()):
        c = por_alpha[a]["count"]
        ws.append([
            f"{a:.1f}",
            por_alpha[a]["val"] / c,
            por_alpha[a]["tempo"] / c,
            int(por_alpha[a]["it"] / c)
        ])

    ajustar_largura_colunas(ws)


def exportar_xlsx(caminho, linhas, tempo_total_script, config):
    wb = Workbook()
    ws = wb.active
    ws.title = "resultados"

    headers = ["heuristica", "n", "m", "replicacao", "tempo", "iteracoes", "valor", "parametro"]
    ws.append(headers)

    for heur, n, m, rep, tempo, it, val, alpha in linhas:
        ws.append([heur, int(n), int(m), int(rep), float(tempo), int(it), int(val), float(alpha)])

    estilizar_cabecalho(ws, num_cols=len(headers))

    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=5).number_format = "0.0000"  # tempo
        ws.cell(row=row, column=8).number_format = "0.0"     # alpha

    ajustar_largura_colunas(ws)

    # Aba extra com resumo
    criar_aba_resumo(wb, linhas, tempo_total_script, config)

    wb.save(caminho)


def main():
    random.seed()

    inicio_script = time.time()

    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    OUT_DIR = os.path.join(BASE_DIR, "Resultados")
    os.makedirs(OUT_DIR, exist_ok=True)
    
    timestamp = time.strftime("%d-%m-%Y_%H-%M-%S")
    
    TXT_PATH = os.path.join(OUT_DIR, f"resultados_blnm_{timestamp}.txt")
    XLSX_PATH = os.path.join(OUT_DIR, f"resultados_blnm_{timestamp}.xlsx")
    
    maquinas = [10, 20, 50]
    rs = [1.5, 2.0]
    repeticoes = 10
    alphas = [i / 10 for i in range(1, 10)]  # 0.1..0.9
    max_sem_melhora = 1000

    linhas = []

    total = len(maquinas) * len(rs) * repeticoes * len(alphas)
    done = 0

    for m in maquinas:
        for r in rs:
            n = int(m * r)

            for rep in range(1, repeticoes + 1):
                tempos = [random.randint(1, 100) for _ in range(n)]

                for alpha in alphas:
                    valor, it, tempo_exec = blnm_monotona_randomizada(
                        tempos, m, alpha, max_sem_melhora=max_sem_melhora
                    )

                    linhas.append((
                        "blnm_monotona_randomizada",
                        n, m, rep,
                        tempo_exec,
                        it,
                        valor,
                        alpha
                    ))

                    done += 1
                    if done % 20 == 0:
                        print(f"[BLNM] {done}/{total} (parcial)")

    tempo_total_script = time.time() - inicio_script

    exportar_txt(TXT_PATH, linhas)

    config = {
        "maquinas": maquinas,
        "rs": rs,
        "repeticoes": repeticoes,
        "alphas": [f"{a:.1f}" for a in alphas],
        "max_sem_melhora": max_sem_melhora,
        "esperado_registros": total
    }

    exportar_xlsx(XLSX_PATH, linhas, tempo_total_script, config)

    print(f"\nGerado:\n- {TXT_PATH}\n- {XLSX_PATH}")
    print(f"Total de registros (esperado {total}): {len(linhas)}")
    print(f"Tempo total do script: {tempo_total_script:.2f}s")


if __name__ == "__main__":
    main()